"""
视频理解测试 - 优化版
目标：输出具体时间点而非时间段
"""
import json
import os
import re
import time
import base64
from pathlib import Path
from collections import defaultdict
import requests

# API配置
API_KEY = "sk-iKGxKJOSZ4ZVAE9noaVSe3ahYkRVP7BbyfqFQtC7x88JatLW"
API_URL = "https://yunwu.ai/v1beta/models/gemini-2.0-flash:generateContent"


def call_gemini(prompt: str, images: list = None) -> str:
    """调用Gemini API"""
    url = f"{API_URL}?key={API_KEY}"
    headers = {"Content-Type": "application/json"}
    
    parts = [{"text": prompt}]
    if images:
        for img in images[:3]:  # 最多3张
            parts.append({"inline_data": {"mime_type": "image/jpeg", "data": img}})
    
    payload = {
        "contents": [{"parts": parts}],
        "generationConfig": {"temperature": 0.7, "maxOutputTokens": 512}
    }
    
    for attempt in range(3):
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=60)
            response.raise_for_status()
            return response.json()['candidates'][0]['content']['parts'][0]['text']
        except Exception as e:
            if attempt < 2:
                time.sleep(3 * (attempt + 1))
            else:
                return "{}"


def load_skill_framework():
    """加载技能框架"""
    skill_file = "data/hangzhou-leiming/skills/ai-drama-clipping-thoughts-v0.1.md"
    with open(skill_file, 'r', encoding='utf-8') as f:
        content = f.read()
    json_start = content.find('```json\n')
    json_end = content.find('\n```', json_start)
    return json.loads(content[json_start+7:json_end])


def load_keyframes(project: str, episode: int) -> list:
    """加载关键帧"""
    base_dir = os.path.expanduser("~/Downloads/hangzhou-leiming-ai-drama/data/hangzhou-leiming/cache/keyframes")
    keyframe_dir = Path(f"{base_dir}/{project}/{episode}")
    if not keyframe_dir.exists():
        return []
    
    frames = sorted(keyframe_dir.glob("*.jpg"))
    result = []
    for i, f in enumerate(frames):
        timestamp_ms = int((i * 500))  # 0.5秒/帧
        result.append({"path": str(f), "timestamp_ms": timestamp_ms})
    return result


def load_asr(project: str, episode: int) -> list:
    """加载ASR"""
    base_dir = os.path.expanduser("~/Downloads/hangzhou-leiming-ai-drama/data/hangzhou-leiming/cache/asr")
    asr_file = Path(f"{base_dir}/{project}/{episode}.json")
    if not asr_file.exists():
        return []
    
    with open(asr_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    segments = []
    for seg in data.get('segments', []):
        segments.append({
            'text': seg.get('text', ''),
            'start': seg.get('start', 0),
            'end': seg.get('end', 0)
        })
    return segments


def load_human_markings(project: str) -> list:
    """加载人工标记"""
    excel_file = f"/Users/wangheng/Downloads/{project}.xlsx"
    if not os.path.exists(excel_file):
        return []
    
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("需要安装openpyxl")
        return []
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    markings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        episode_str = str(row[0])
        time_str = str(row[1])
        mark_type = str(row[2])
        
        match = re.search(r'第(\d+)集', episode_str)
        if match:
            episode = int(match.group(1))
            parts = time_str.split(':')
            if len(parts) >= 2:
                seconds = int(parts[0]) * 60 + int(parts[1])
            else:
                seconds = 0
            
            markings.append({
                'episode': episode,
                'timestamp': seconds,
                'type': mark_type
            })
    
    return markings


def find_key_moments(keyframes: list, asr_segments: list, skill_framework: dict) -> dict:
    """
    优化版：直接找出具体时间点
    
    策略：
    1. 分析每个ASR片段的边界（可能的时间点）
    2. 结合关键帧，让AI判断哪些是具体的高光/钩子时间点
    3. 输出具体秒数
    """
    results = {
        'highlights': [],  # 具体时间点（秒）
        'hooks': []       # 具体时间点（秒）
    }
    
    if not asr_segments:
        return results
    
    # 找到每个场景切换点（ASR片段边界）
    potential_points = []
    for seg in asr_segments:
        # 记录片段开始时间作为潜在点
        potential_points.append({
            'time': int(seg['start']),
            'text': seg['text'][:100],
            'type': 'asr_boundary'
        })
    
    # 采样关键帧时间点（每10秒一个）
    for i, kf in enumerate(keyframes):
        if i % 20 == 0:  # 每10秒
            potential_points.append({
                'time': kf['timestamp_ms'] // 1000,
                'text': '',
                'type': 'keyframe'
            })
    
    # 去重并排序
    unique_times = sorted(list(set([p['time'] for p in potential_points])))
    
    # 对每个时间点进行判断（但不要全部判断，太慢）
    # 选择关键时间点判断
    sample_times = unique_times[::max(1, len(unique_times)//15)]  # 采样约15个点
    
    for t in sample_times:
        # 获取该时间点前后的上下文
        context_asr = [s for s in asr_segments if t-10 < s['end'] and s['start'] < t+10]
        context_text = ' '.join([s['text'] for s in context_asr])[:300]
        
        # 获取附近关键帧
        context_kf = [kf for kf in keyframes if abs(kf['timestamp_ms']/1000 - t) < 10]
        
        # 构建prompt - 要求输出具体时间点
        prompt = f"""分析这个短剧时间点 ({t}秒附近)：
        
上下文ASR：{context_text}

根据以下钩子类型定义：
{json.dumps(skill_framework.get('hook_types', [])[:5], ensure_ascii=False, indent=2)}

判断这个时间点是否是"钩子点"（让人想继续看下去的关键时刻）。
注意：只需要判断这1个具体时间点。

返回JSON格式：
{{
  "isHook": true/false,
  "hookTime": {t},  // 具体秒数
  "hookType": "类型名或null",
  "reason": "简短原因"
}}

只返回JSON。"""

        try:
            imgs = []
            for kf in context_kf[:2]:
                try:
                    with open(kf['path'], 'rb') as f:
                        imgs.append(base64.b64encode(f.read()).decode('utf-8'))
                except:
                    pass
            
            response = call_gemini(prompt, imgs)
            match = re.search(r'\{[\s\S]*\}', response)
            if match:
                result = json.loads(match.group())
                if result.get('isHook') and result.get('hookTime'):
                    results['hooks'].append({
                        'time': result.get('hookTime'),
                        'type': result.get('hookType'),
                        'reason': result.get('reason', '')
                    })
        except Exception as e:
            print(f"  错误: {e}")
    
    return results


def analyze_video_direct(keyframes: list, asr_segments: list, skill_framework: dict) -> dict:
    """
    直接分析模式：让AI直接输出时间点列表
    不分段，直接看整体然后输出时间点
    """
    results = {
        'highlights': [],
        'hooks': []
    }
    
    if not asr_segments:
        return results
    
    # 取前60秒的ASR作为样本
    sample_asr = [s for s in asr_segments if s['end'] < 60]
    sample_text = ' '.join([s['text'] for s in sample_asr])[:500]
    
    # 获取样本关键帧
    sample_kf = keyframes[:6]
    
    imgs = []
    for kf in sample_kf:
        try:
            with open(kf['path'], 'rb') as f:
                imgs.append(base64.b64encode(f.read()).decode('utf-8'))
        except:
            pass
    
    # 让AI直接给出具体时间点
    prompt = f"""你是一个短剧剪辑专家。

根据以下视频开头（0-60秒）的内容：
ASR文本：{sample_text}

参考钩子类型定义：
{json.dumps(skill_framework.get('hook_types', [])[:5], ensure_ascii=False, indent=2)}

请直接列出这个片段中所有具体的"钩子点"时间（秒），格式如下：
- 只列出具体秒数
- 例如：35, 120, 185
- 如果没有明确的钩子点，回复"无"

只回复时间列表，不要其他文字。"""

    try:
        response = call_gemini(prompt, imgs)
        print(f"  AI响应: {response[:200]}")
        
        # 提取时间
        times = re.findall(r'(\d+)', response)
        for t in times:
            results['hooks'].append({
                'time': int(t),
                'type': '钩子点',
                'reason': 'AI直接识别'
            })
    except Exception as e:
        print(f"  错误: {e}")
    
    return results


def main():
    project = "百里将就"
    
    print("=" * 60)
    print("视频理解测试 - 优化版（输出具体时间点）")
    print("=" * 60)
    
    # 1. 加载数据
    print("\n[1] 加载数据...")
    skill_data = load_skill_framework()
    human_markings = load_human_markings(project)
    
    print(f"  技能: {len(skill_data['highlight_types'])}高光类型, {len(skill_data['hook_types'])}钩子类型")
    print(f"  人工标记: {len(human_markings)}个")
    
    # 2. 分析第1集作为测试
    episode = 1
    print(f"\n[2] 分析第{episode}集...")
    
    keyframes = load_keyframes(project, episode)
    asr_segments = load_asr(project, episode)
    
    print(f"  关键帧: {len(keyframes)}")
    print(f"  ASR段: {len(asr_segments)}")
    
    if not asr_segments:
        print("  错误：无ASR数据")
        return
    
    # 3. 直接分析模式（让AI输出具体时间点）
    print("\n[3] 直接分析模式...")
    results = analyze_video_direct(keyframes, asr_segments, skill_data)
    
    print(f"\n  AI识别钩子点: {len(results['hooks'])}个")
    for h in results['hooks']:
        print(f"    {h['time']}秒 - {h['type']}")
    
    # 4. 对比
    print("\n[4] 对比...")
    human_hooks = [m for m in human_markings if '钩子' in m['type'] and m['episode'] == episode]
    
    print(f"  人工标记: {len(human_hooks)}个")
    for h in human_hooks:
        print(f"    {h['timestamp']}秒")
    
    # 匹配
    matched = 0
    for hh in human_hooks:
        for ah in results['hooks']:
            if abs(hh['timestamp'] - ah['time']) <= 10:  # 10秒内算匹配
                matched += 1
                print(f"  ✓ 匹配: 人工{hh['timestamp']}秒 ≈ AI{ah['time']}秒")
                break
    
    print(f"\n  匹配: {matched}/{len(human_hooks)}")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
