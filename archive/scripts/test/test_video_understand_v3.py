"""
视频理解测试 - 优化版v2
分析完整视频，输出具体时间点
"""
import json
import os
import re
import time
import base64
from pathlib import Path
import requests

API_KEY = "sk-iKGxKJOSZ4ZVAE9noaVSe3ahYkRVP7BbyfqFQtC7x88JatLW"
API_URL = "https://yunwu.ai/v1beta/models/gemini-2.0-flash:generateContent"


def call_gemini(prompt: str, images: list = None) -> str:
    url = f"{API_URL}?key={API_KEY}"
    parts = [{"text": prompt}]
    if images:
        for img in images[:3]:
            parts.append({"inline_data": {"mime_type": "image/jpeg", "data": img}})
    
    for attempt in range(3):
        try:
            r = requests.post(url, headers={"Content-Type": "application/json"},
                json={"contents": [{"parts": parts}], "generationConfig": {"temperature": 0.7, "maxOutputTokens": 256}}, timeout=60)
            return r.json()['candidates'][0]['content']['parts'][0]['text']
        except:
            if attempt < 2: time.sleep(3)
    return "{}"


def load_skill():
    with open("data/hangzhou-leiming/skills/ai-drama-clipping-thoughts-v0.1.md") as f:
        content = f.read()
    return json.loads(content[content.find('```json')+7:content.find('\n```')])


def load_asr(project, ep):
    path = Path(f"~/Downloads/hangzhou-leiming-ai-drama/data/hangzhou-leiming/cache/asr/{project}/{ep}.json")
    if not path.exists(): return []
    with open(path) as f:
        return [{'text': s['text'], 'start': s['start'], 'end': s['end']} for s in json.load(f)['segments']]


def load_kf(project, ep):
    d = Path(f"~/Downloads/hangzhou-leiming-ai-drama/data/hangzhou-leiming/cache/keyframes/{project}/{ep}")
    if not d.exists(): return []
    return sorted(d.glob("*.jpg"))


def load_human(project):
    from openpyxl import load_workbook
    wb = load_workbook(f"/Users/wangheng/Downloads/{project}.xlsx")
    result = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            m = re.search(r'第(\d+)集', str(row[0]))
            if m:
                ts = str(row[1]).split(':')
                result.append({'ep': int(m.group(1)), 'ts': int(ts[0])*60+int(ts[1]), 'type': str(row[2])})
    return result


def analyze_segment(segment_asr, segment_kf, skill, segment_name):
    """分析单个时间段的ASR，让AI找出该段内的钩子点"""
    if not segment_asr:
        return []
    
    text = ' '.join([s['text'] for s in segment_asr])[:600]
    
    imgs = []
    for p in segment_kf[:2]:
        try:
            with open(p, 'rb') as f:
                imgs.append(base64.b64encode(f.read()).decode())
        except: pass
    
    # 让AI列出该段内的具体钩子点时间
    prompt = f"""分析这段短剧内容：

ASR文本：{text}

参考钩子类型：
{json.dumps(skill.get('hook_types', [])[:5], ensure_ascii=False)}

找出这段中所有具体的"钩子点"时间（秒），只列出数字，用逗号分隔。
例如：35,120,185
如果无钩子点，回复"无"。

只回复数字列表。"""

    try:
        resp = call_gemini(prompt, imgs)
        # 提取数字
        times = re.findall(r'\d+', resp)
        return [int(t) for t in times if int(t) < 500]  # 过滤异常大的数字
    except:
        return []


def main():
    project = "百里将就"
    
    print("=" * 60)
    print("视频理解测试 - 分析完整视频")
    print("=" * 60)
    
    skill = load_skill()
    human = load_human(project)
    
    print(f"\n人工标记: {len(human)}个")
    human_hooks = [h for h in human if '钩子' in h['type']]
    print(f"其中钩子点: {len(human_hooks)}个")
    
    all_results = []
    
    # 分析每集
    for ep in [1,2,3,4,6,7]:
        print(f"\n=== 第{ep}集 ===")
        
        asr = load_asr(project, ep)
        kf_files = load_kf(project, ep)
        
        if not asr:
            print(f"  无ASR数据")
            continue
        
        duration = int(asr[-1]['end'])
        print(f"  时长: {duration}秒, ASR段: {len(asr)}")
        
        # 分段分析（每30秒一段，覆盖整集）
        segment_size = 30
        hooks_in_ep = []
        
        for start in range(0, duration, segment_size):
            end = min(start + segment_size, duration)
            
            # 该段的ASR
            seg_asr = [s for s in asr if s['start'] < end and s['end'] > start]
            if not seg_asr:
                continue
            
            # 该段的关键帧
            seg_kf = [kf_files[i] for i in range(len(kf_files)) if start*2 <= i*500 < end*2][::10]
            
            # 分析
            times = analyze_segment(seg_asr, seg_kf, skill, f"{start}-{end}")
            
            for t in times:
                # 转换为绝对时间
                abs_time = start + t
                if abs_time <= duration:
                    hooks_in_ep.append(abs_time)
                    print(f"  {start}-{end}秒段内: {t}秒 -> 绝对时间{abs_time}秒")
        
        all_results.append({'ep': ep, 'hooks': hooks_in_ep})
        print(f"  第{ep}集识别到 {len(hooks_in_ep)} 个钩子点")
    
    # 对比
    print("\n" + "=" * 60)
    print("对比结果")
    print("=" * 60)
    
    total_ai = 0
    total_human = len(human_hooks)
    matched = 0
    
    for ep_data in all_results:
        ep = ep_data['ep']
        ai_hooks = ep_data['hooks']
        human_in_ep = [h for h in human_hooks if h['ep'] == ep]
        
        print(f"\n第{ep}集:")
        print(f"  人工: {[h['ts'] for h in human_in_ep]}")
        print(f"  AI:   {ai_hooks}")
        
        for hh in human_in_ep:
            for ah in ai_hooks:
                if abs(hh['ts'] - ah) <= 10:
                    matched += 1
                    print(f"    ✓ {hh['ts']}秒 ≈ {ah}秒")
                    break
        
        total_ai += len(ai_hooks)
    
    print(f"\n总计:")
    print(f"  AI识别: {total_ai}个")
    print(f"  人工标记: {total_human}个")
    print(f"  匹配: {matched}/{total_human} = {matched*100//total_human if total_human else 0}%")


if __name__ == "__main__":
    main()
