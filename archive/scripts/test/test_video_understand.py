"""
完整的视频理解测试 - AI分析 + 对比报告
"""
import json
import os
import re
import sys
import time
import base64
from pathlib import Path
from collections import defaultdict
import requests

# API配置
API_KEY = "sk-iKGxKJOSZ4ZVAE9noaVSe3ahYkRVP7BbyfqFQtC7x88JatLW"
API_URL = "https://yunwu.ai/v1beta/models/gemini-2.0-flash:generateContent"

# 分析配置
SEGMENT_DURATION = 60  # 片段时长（秒）
SEGMENT_OVERLAP = 10   # 重叠时长（秒）
FRAMES_PER_SEGMENT = 3 # 每片段关键帧数


def call_gemini(prompt: str, images: list = None) -> str:
    """调用Gemini API"""
    url = f"{API_URL}?key={API_KEY}"
    headers = {"Content-Type": "application/json"}
    
    parts = [{"text": prompt}]
    if images:
        for img in images[:5]:  # 最多5张
            parts.append({"inline_data": {"mime_type": "image/jpeg", "data": img}})
    
    payload = {
        "contents": [{"parts": parts}],
        "generationConfig": {"temperature": 0.7, "maxOutputTokens": 512}
    }
    
    for attempt in range(3):
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=60)
            response.raise_for_status()
            return response.json()['candidates'][0]['content']['parts'][0]['text']
        except Exception as e:
            if attempt < 2:
                time.sleep(5 * (attempt + 1))
            else:
                raise e


def load_skill_framework():
    """加载技能框架"""
    skill_file = "data/hangzhou-leiming/skills/ai-drama-clipping-thoughts-v0.1.md"
    with open(skill_file, 'r', encoding='utf-8') as f:
        content = f.read()
    json_start = content.find('```json\n')
    json_end = content.find('\n```', json_start)
    skill_data = json.loads(content[json_start+7:json_end])
    return skill_data


def load_keyframes(project: str, episode: int) -> list:
    """加载关键帧"""
    base_dir = os.path.expanduser("~/Downloads/hangzhou-leiming-ai-drama/data/hangzhou-leiming/cache/keyframes")
    keyframe_dir = Path(f"{base_dir}/{project}/{episode}")
    if not keyframe_dir.exists():
        return []
    
    frames = sorted(keyframe_dir.glob("*.jpg"))
    result = []
    for i, f in enumerate(frames):
        timestamp_ms = int((i * 500))  # 0.5秒/帧
        result.append({"path": str(f), "timestamp_ms": timestamp_ms})
    return result


def load_asr(project: str, episode: int) -> list:
    """加载ASR"""
    base_dir = os.path.expanduser("~/Downloads/hangzhou-leiming-ai-drama/data/hangzhou-leiming/cache/asr")
    asr_file = Path(f"{base_dir}/{project}/{episode}.json")
    if not asr_file.exists():
        return []
    
    with open(asr_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    segments = []
    for seg in data.get('segments', []):
        segments.append({
            'text': seg.get('text', ''),
            'start': seg.get('start', 0),
            'end': seg.get('end', 0)
        })
    return segments


def load_human_markings(project: str) -> list:
    """加载人工标记"""
    excel_file = f"/Users/wangheng/Downloads/{project}.xlsx"
    if not os.path.exists(excel_file):
        return []
    
    try:
        from openpyxl import load_workbook
    except:
        print("需要安装openpyxl: pip install openpyxl")
        return []
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    markings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        episode_str = str(row[0])
        time_str = str(row[1])
        mark_type = str(row[2])
        
        match = re.search(r'第(\d+)集', episode_str)
        if match:
            episode = int(match.group(1))
            parts = time_str.split(':')
            if len(parts) >= 2:
                seconds = int(parts[0]) * 60 + int(parts[1])
            else:
                seconds = 0
            
            markings.append({
                'episode': episode,
                'timestamp': seconds,
                'type': mark_type
            })
    
    return markings


def extract_segments(keyframes: list, asr_segments: list, video_duration: int) -> list:
    """提取分析片段"""
    segments = []
    start = 0
    
    while start < video_duration:
        end = min(start + SEGMENT_DURATION, video_duration)
        
        # 筛选关键帧
        seg_kf = [kf for kf in keyframes if start*1000 <= kf['timestamp_ms'] < end*1000]
        # 均匀采样
        if len(seg_kf) > FRAMES_PER_SEGMENT:
            indices = [int(i * len(seg_kf) / FRAMES_PER_SEGMENT) for i in range(FRAMES_PER_SEGMENT)]
            seg_kf = [seg_kf[i] for i in indices]
        
        # 筛选ASR
        seg_asr = [s for s in asr_segments if start < s['end'] and s['start'] < end]
        
        segments.append({
            'start': start,
            'end': end,
            'keyframes': seg_kf,
            'asr': seg_asr
        })
        
        start = start + SEGMENT_DURATION - SEGMENT_OVERLAP
    
    return segments


def analyze_segment(segment: dict, skill_framework: dict) -> dict:
    """分析单个片段"""
    # 加载关键帧图片
    images = []
    for kf in segment['keyframes']:
        try:
            with open(kf['path'], 'rb') as f:
                images.append(base64.b64encode(f.read()).decode('utf-8'))
        except:
            pass
    
    # ASR文本
    asr_text = ' '.join([s['text'] for s in segment['asr']])[:500]
    
    highlight_types = skill_framework.get('highlight_types', [])
    hook_types = skill_framework.get('hook_types', [])
    
    prompt = f"""你是一位短剧剪辑专家。请分析以下视频片段：

时间范围：{segment['start']}-{segment['end']}秒
关键帧：{len(images)}张
ASR文本：{asr_text[:300]}

高光类型定义：{json.dumps(highlight_types[:5], ensure_ascii=False, indent=2)}
钩子类型定义：{json.dumps(hook_types[:5], ensure_ascii=False, indent=2)}

请判断这个片段是否包含：
1. 高光点（精彩/重要内容）？
2. 钩子点（让人想继续看）？

返回JSON格式（直接返回，不要其他文字）：
{{
  "isHighlight": true/false,
  "highlightType": "类型名或null",
  "highlightDesc": "简短描述",
  "isHook": true/false, 
  "hookType": "类型名或null",
  "hookDesc": "简短描述"
}}"""

    try:
        response = call_gemini(prompt, images)
        match = re.search(r'\{[\s\S]*\}', response)
        if match:
            return json.loads(match.group())
    except Exception as e:
        print(f"  API错误: {e}")
    
    return {
        "isHighlight": False,
        "highlightType": None,
        "highlightDesc": "",
        "isHook": False,
        "hookType": None,
        "hookDesc": ""
    }


def main():
    project = "百里将就"
    import sys
    
    print("=" * 60, flush=True)
    print("视频理解测试 - AI分析与人工标记对比", flush=True)
    print("=" * 60, flush=True)
    
    # 1. 加载数据
    print("\n[1] 加载数据...")
    skill_data = load_skill_framework()
    human_markings = load_human_markings(project)
    
    print(f"  技能: {len(skill_data['highlight_types'])}高光类型, {len(skill_data['hook_types'])}钩子类型")
    print(f"  人工标记: {len(human_markings)}个")
    
    # 2. 获取有数据的集数
    episodes_with_data = []
    for ep in [1, 2, 3, 4, 6, 7]:
        kf = load_keyframes(project, ep)
        asr = load_asr(project, ep)
        if kf or asr:
            episodes_with_data.append(ep)
    
    print(f"  有数据的集数: {episodes_with_data}")
    
    # 3. AI分析
    print("\n[2] AI分析中...")
    all_ai_results = []
    total_segments = 0
    
    for episode in episodes_with_data:
        print(f"\n  分析第{episode}集...")
        keyframes = load_keyframes(project, episode)
        asr_segments = load_asr(project, episode)
        
        if not keyframes:
            continue
        
        # 计算视频时长
        video_duration = (keyframes[-1]['timestamp_ms'] // 1000) + 10
        
        # 提取片段
        segments = extract_segments(keyframes, asr_segments, video_duration)
        total_segments += len(segments)
        print(f"    {len(segments)}个片段")
        
        # 分析每个片段
        for i, seg in enumerate(segments):
            result = analyze_segment(seg, skill_data)
            result['episode'] = episode
            result['start_time'] = seg['start']
            result['end_time'] = seg['end']
            
            if result['isHighlight'] or result['isHook']:
                marker = []
                if result['isHighlight']:
                    marker.append(f"高光:{result['highlightType']}")
                if result['isHook']:
                    marker.append(f"钩子:{result['hookType']}")
                print(f"    片段{i+1}: {seg['start']}-{seg['end']}秒 -> {', '.join(marker)}")
            
            all_ai_results.append(result)
    
    # 4. 整理AI结果
    ai_highlights = [r for r in all_ai_results if r['isHighlight']]
    ai_hooks = [r for r in all_ai_results if r['isHook']]
    
    print(f"\n  总片段: {total_segments}")
    print(f"  AI识别高光点: {len(ai_highlights)}个")
    print(f"  AI识别钩子点: {len(ai_hooks)}个")
    
    # 5. 与人工标记对比
    print("\n[3] 对比分析...")
    
    # 人工标记
    human_hooks = [m for m in human_markings if '钩子' in m['type']]
    print(f"  人工标记钩子点: {len(human_hooks)}个")
    
    # 匹配检测（30秒内算匹配）
    matched = 0
    for hh in human_hooks:
        for ah in ai_hooks:
            if hh['episode'] == ah['episode']:
                diff = abs(hh['timestamp'] - ah['start_time'])
                if diff <= 30:
                    matched += 1
                    print(f"    匹配: 第{hh['episode']}集 人工{hh['timestamp']}秒 ≈ AI{ah['start_time']}秒 ({ah['hookType']})")
                    break
    
    print(f"  匹配数: {matched}/{len(human_hooks)}")
    
    # 6. 详细对比表
    print("\n[4] 详细对比表")
    print("\n  人工标记的钩子点:")
    for m in human_hooks:
        print(f"    第{m['episode']}集 {m['timestamp']}秒")
    
    print("\n  AI识别的钩子点(前20个):")
    for r in ai_hooks[:20]:
        print(f"    第{r['episode']}集 {r['start_time']}-{r['end_time']}秒 - {r['hookType']}")
    
    # 7. 保存结果
    result = {
        "project": project,
        "total_segments": total_segments,
        "ai_highlights": len(ai_highlights),
        "ai_hooks": len(ai_hooks),
        "human_hooks": len(human_hooks),
        "matched": matched,
        "match_rate": f"{matched/len(human_hooks)*100:.1f}%" if human_hooks else "N/A",
        "ai_results": all_ai_results
    }
    
    output_file = f"data/hangzhou-leiming/analysis/{project}_result.json"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\n结果已保存: {output_file}")
    
    print("\n" + "=" * 60)
    print("测试完成!")
    print("=" * 60)


if __name__ == "__main__":
    main()
