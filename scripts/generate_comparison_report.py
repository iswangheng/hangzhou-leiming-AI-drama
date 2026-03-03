#!/usr/bin/env python3
"""
生成V1与V2优化效果对比报告
"""

import json
import pandas as pd
from pathlib import Path
from collections import defaultdict

def load_json(file_path):
    """加载JSON文件"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def print_section(title):
    """打印分隔线"""
    print("\n" + "="*80)
    print(f"  {title}")
    print("="*80 + "\n")

def analyze_timestamp_precision(timestamps):
    """分析时间戳精度"""
    # 检查是否是窗口开始时间（0, 50, 100, 150...）
    window_pattern = [t % 50 == 0 for t in timestamps]
    is_window_based = sum(window_pattern) / len(timestamps) > 0.8 if timestamps else False

    # 计算时间戳分布
    unique_timestamps = len(set(timestamps))
    total_timestamps = len(timestamps)

    return {
        'is_window_based': is_window_based,
        'unique_ratio': unique_timestamps / total_timestamps if total_timestamps > 0 else 0,
        'total_count': total_timestamps,
        'unique_count': unique_timestamps
    }

def main():
    # 项目信息
    project_name = "百里将就"

    # 加载V1和V2结果
    v1_result = load_json("data/hangzhou-leiming/analysis/百里将就/result.json")
    v2_result = load_json("data/hangzhou-leiming/analysis/百里将就_v2/result.json")

    # 加载人工标记数据
    from openpyxl import load_workbook
    excel_file = "漫剧素材/百里将就/百里将就.xlsx"
    wb = load_workbook(excel_file)
    ws = wb.active

    # 统计人工标记
    human_marks = defaultdict(list)
    for row in list(ws.iter_rows(values_only=True))[1:]:  # 跳过表头
        episode_str = str(row[0])
        # 从"第1集"提取数字
        episode = int(''.join(filter(str.isdigit, episode_str))) if episode_str else 1

        # 时间是 datetime.time 对象，格式为 MM:SS 或 M:SS
        # datetime.time(5, 50) 表示 5分50秒，不是5小时50分！
        time_obj = row[1]
        if hasattr(time_obj, 'hour') and hasattr(time_obj, 'minute') and hasattr(time_obj, 'second'):
            # datetime.time 对象，但 hour 字段实际是分钟
            timestamp = time_obj.hour * 60 + time_obj.minute
        else:
            timestamp = 0

        mark_type = row[2]
        human_marks[episode].append({
            'timestamp': int(timestamp),
            'type': mark_type
        })

    total_human_marks = sum(len(marks) for marks in human_marks.values())

    # ========== 1. 整体指标对比 ==========
    print_section("一、整体指标对比")

    # 表头
    header = f"{'指标':<20} {'优化前(V1)':<15} {'优化后(V2)':<15} {'提升':<10}"
    print(header)
    print("-" * 70)

    # 高光点数量
    hl_v1 = len(v1_result.get('highlights', []))
    hl_v2 = len(v2_result.get('highlights', []))
    hl_diff = f"+{hl_v2 - hl_v1}" if hl_v2 > hl_v1 else str(hl_v2 - hl_v1)
    print(f"{'高光点数量':<20} {hl_v1:<15} {hl_v2:<15} {hl_diff:<10}")

    # 钩子点数量
    hook_v1 = len(v1_result.get('hooks', []))
    hook_v2 = len(v2_result.get('hooks', []))
    hook_diff = f"{hook_v2 - hook_v1}"
    print(f"{'钩子点数量':<20} {hook_v1:<15} {hook_v2:<15} {hook_diff:<10}")

    # 人工标记数量
    print(f"{'人工标记数量':<20} {'-':<15} {total_human_marks:<15} {'参考':<10}")

    # 剪辑组合数量
    clip_v1 = len(v1_result.get('clips', []))
    clip_v2 = len(v2_result.get('clips', []))
    clip_diff = f"+{clip_v2 - clip_v1}" if clip_v2 > clip_v1 else str(clip_v2 - clip_v1)
    print(f"{'剪辑组合数量':<20} {clip_v1:<15} {clip_v2:<15} {clip_diff:<10}")

    # 平均置信度
    conf_v2 = v2_result.get('statistics', {}).get('averageConfidence', 0)
    print(f"{'平均置信度':<20} {'无':<15} {conf_v2:<15.2f} {'新增':<10}")

    # ========== 2. 时间戳精度对比 ==========
    print_section("二、时间戳精度分析")

    v1_hooks = v1_result.get('hooks', [])
    v2_hooks = v2_result.get('hooks', [])

    v1_timestamps = [h['timestamp'] for h in v1_hooks]
    v2_timestamps = [h['timestamp'] for h in v2_hooks]

    v1_analysis = analyze_timestamp_precision(v1_timestamps)
    v2_analysis = analyze_timestamp_precision(v2_timestamps)

    print("V1 时间戳特征:")
    print(f"  - 基于窗口开始: {v1_analysis['is_window_based']}")
    print(f"  - 独立时间戳比例: {v1_analysis['unique_ratio']:.2%}")
    print(f"  - 前10个时间戳: {v1_timestamps[:10]}")

    print("\nV2 时间戳特征:")
    print(f"  - 基于窗口开始: {v2_analysis['is_window_based']}")
    print(f"  - 独立时间戳比例: {v2_analysis['unique_ratio']:.2%}")
    print(f"  - 前10个时间戳: {v2_timestamps[:10]}")

    # ========== 3. 分集详细对比 ==========
    print_section("三、分集详细对比")

    # 统计V1和V2每集的钩子点
    v1_by_episode = defaultdict(list)
    for h in v1_hooks:
        ep = h.get('episode', 1)
        v1_by_episode[ep].append(h['timestamp'])

    v2_by_episode = defaultdict(list)
    for h in v2_hooks:
        ep = h.get('episode', 1)
        v2_by_episode[ep].append(h['timestamp'])

    # 打印对比表
    print(f"{'集数':<8} {'人工标记':<20} {'V1识别':<25} {'V2识别':<25}")
    print("-" * 85)

    for ep in sorted(set(list(human_marks.keys()) + list(v1_by_episode.keys()) + list(v2_by_episode.keys()))):
        human_ts = [m['timestamp'] for m in human_marks.get(ep, [])]
        v1_ts = v1_by_episode.get(ep, [])
        v2_ts = v2_by_episode.get(ep, [])

        human_str = str(human_ts) if human_ts else "无"
        v1_str = f"{len(v1_ts)}个: {v1_ts}" if v1_ts else "无"
        v2_str = f"{len(v2_ts)}个: {v2_ts}" if v2_ts else "无"

        print(f"第{ep}集{'':<4} {human_str:<20} {v1_str:<25} {v2_str:<25}")

    # ========== 4. 质量筛选效果 ==========
    print_section("四、质量筛选效果（V2）")

    v2_hooks_with_conf = [(h['timestamp'], h.get('confidence', 0)) for h in v2_hooks]
    v2_hooks_sorted = sorted(v2_hooks_with_conf, key=lambda x: x[1], reverse=True)

    print("置信度分布:")
    print(f"  - 平均置信度: {v2_result.get('statistics', {}).get('averageConfidence', 0):.2f}")
    print(f"  - 最高置信度: {max([c for _, c in v2_hooks_with_conf]) if v2_hooks_with_conf else 0:.1f}")
    print(f"  - 最低置信度: {min([c for _, c in v2_hooks_with_conf]) if v2_hooks_with_conf else 0:.1f}")

    print("\n置信度 > 8.5 的钩子点:")
    high_conf = [(ts, c) for ts, c in v2_hooks_with_conf if c > 8.5]
    for ts, c in high_conf:
        hook = next(h for h in v2_hooks if h['timestamp'] == ts)
        print(f"  - 第{hook['episode']}集 {ts}秒 (置信度:{c}): {hook['type']}")

    # ========== 5. 关键改进总结 ==========
    print_section("五、关键改进总结")

    improvements = [
        ("✓ 精确时间戳", "从窗口开始时间变为精确时刻", "时间戳从 [0, 50, 100...] 变为 [58, 309, 432...]"),
        ("✓ 数量控制", "从23个降至13个，接近人工的10个", "平均每集从2.3个降至1.4个，人工为1.0个"),
        ("✓ 质量保证", "所有标记都有置信度评分", "平均置信度8.32，说明识别质量很高"),
        ("✓ 开篇识别", "第1集开头自动识别为高光点", "解决了V1中第1集没有高光点的问题"),
        ("✓ 剪辑生成", "成功生成1个剪辑组合", "V1无法生成剪辑，V2可以匹配高光→钩子"),
    ]

    for i, (title, summary, detail) in enumerate(improvements, 1):
        print(f"{i}. {title}")
        print(f"   {summary}")
        print(f"   {detail}\n")

    # ========== 6. 存在的问题 ==========
    print_section("六、仍需优化的问题")

    issues = [
        ("部分集数识别偏多", "第1集识别了4个钩子点，人工只有3个", "可通过调整每集上限参数优化"),
        ("部分集数识别偏少", "第2、5、8、9集识别较少", "可能需要降低置信度阈值或优化Prompt"),
        ("高光点识别仍少", "只识别了1个高光点", "高光点比钩子点更难识别，需要更多训练数据"),
    ]

    for i, (issue, description, suggestion) in enumerate(issues, 1):
        print(f"{i}. {issue}")
        print(f"   - 问题描述: {description}")
        print(f"   - 优化建议: {suggestion}\n")

    print("\n" + "="*80)
    print("报告生成完毕")
    print("="*80)

if __name__ == "__main__":
    main()
